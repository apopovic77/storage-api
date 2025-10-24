"""
AI Analysis Service

Central service for all AI content analysis.
Supports UNIFIED and SPLIT modes.
"""

import httpx
import os
import base64
import json
import io
from typing import Dict, Any, Optional, List

from ai_analysis.config import ai_config, AIAnalysisMode
from ai_analysis.prompts import (
    UNIFIED_PROMPT,
    SAFETY_PROMPT,
    EMBEDDING_PROMPT,
    CHUNKED_CSV_PROMPT,
    VISION_ANALYSIS_PROMPT,
    build_context_info
)
import asyncio

# API configuration
INTERNAL_API_KEY = os.getenv("INTERNAL_API_KEY", "Inetpass1")
API_BASE_URL = os.getenv("AI_API_URL", "https://api-ai.arkturian.com")  # AI microservice

# CSV Chunking configuration
CSV_CHUNK_SIZE = 10  # Process 10 rows per chunk (small for testing/debugging)
CSV_MAX_ROWS_BEFORE_CHUNKING = 50  # Chunk if CSV has more than 50 rows


def extract_excel_as_text(data: bytes, mime_type: str) -> str:
    """
    Extract text from Excel files (.xlsx, .xls) as CSV-like format.

    Handles multiple sheets by extracting all sheets and formatting them
    as separate sections in the output text.

    Args:
        data: Excel file content as bytes
        mime_type: MIME type of the file

    Returns:
        Formatted text representation of all sheets in CSV-like format
    """
    try:
        import openpyxl
        from openpyxl import load_workbook

        # Load workbook from bytes
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

        result_parts = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Add sheet header
            result_parts.append(f"\n=== SHEET: {sheet_name} ===\n")

            # Extract rows as CSV
            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Filter out completely empty rows
                if any(cell is not None and str(cell).strip() for cell in row):
                    # Convert row to strings, handling None values
                    row_strings = [str(cell) if cell is not None else "" for cell in row]
                    rows.append(",".join(row_strings))

            result_parts.append("\n".join(rows))

            # Limit to prevent huge outputs (first 100 rows per sheet)
            if len(rows) > 100:
                result_parts.append(f"\n... ({len(rows) - 100} more rows)")
                break

        wb.close()

        return "\n".join(result_parts)

    except Exception as e:
        # Fallback: try xlrd for older .xls files
        try:
            import xlrd

            wb = xlrd.open_workbook(file_contents=data)
            result_parts = []

            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                sheet_name = sheet.name

                result_parts.append(f"\n=== SHEET: {sheet_name} ===\n")

                rows = []
                for row_idx in range(min(sheet.nrows, 100)):  # Limit to 100 rows
                    row = sheet.row_values(row_idx)
                    row_strings = [str(cell) if cell is not None else "" for cell in row]
                    rows.append(",".join(row_strings))

                result_parts.append("\n".join(rows))

                if sheet.nrows > 100:
                    result_parts.append(f"\n... ({sheet.nrows - 100} more rows)")

            return "\n".join(result_parts)

        except Exception as e2:
            raise Exception(f"Failed to parse Excel file: openpyxl error: {e}, xlrd error: {e2}")


async def _analyze_csv_chunk(
    chunk_text: str,
    chunk_index: int,
    context_info: str
) -> Dict[str, Any]:
    """
    Analyze a single CSV chunk.

    Args:
        chunk_text: CSV text for this chunk (includes header)
        chunk_index: Index of this chunk
        context_info: Context information string

    Returns:
        Dict with embeddingsList and quality_score
    """
    prompt = CHUNKED_CSV_PROMPT.format(
        context_info=context_info,
        chunk_index=chunk_index + 1
    )
    prompt += f"\n\n--- CSV CHUNK ---\n{chunk_text}"

    payload = {"prompt": {"text": prompt, "images": []}}
    headers = {"X-API-KEY": INTERNAL_API_KEY, "Content-Type": "application/json"}

    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{API_BASE_URL}/ai/gemini",
                json=payload,
                headers=headers,
                timeout=120.0  # 2 minutes per chunk
            )
            response.raise_for_status()

            ai_response_str = response.json().get("message", "{}")
            ai_response_str = _clean_json_response(ai_response_str)
            result = json.loads(ai_response_str)

            return {
                "embeddingsList": result.get("embeddingsList", []),
                "quality_score": result.get("quality_score", 5)
            }
        except Exception as e:
            print(f"❌ Chunk {chunk_index + 1} failed: {e}")
            return {
                "embeddingsList": [],
                "quality_score": 0
            }


async def _analyze_csv_chunked(
    csv_text: str,
    context: Optional[Dict[str, Any]] = None,
    object_id: Optional[int] = None
) -> Dict[str, Any]:
    """
    Process large CSV in chunks with progress tracking.

    Args:
        csv_text: Full CSV text content
        context: Optional contextual metadata
        object_id: Optional object ID for progress tracking

    Returns:
        Dict with merged results from all chunks
    """
    import time
    import json
    from pathlib import Path

    context_info = build_context_info(context)
    started_at = time.time()

    # Progress tracking
    def write_progress(stage, chunk_idx=None, details=None):
        if not object_id:
            return
        try:
            status_dir = Path(f"/tmp/ai_analysis_status_{object_id}")
            status_file = status_dir / "progress.json"
            progress = {
                "status": "processing",
                "stage": stage,
                "started_at": started_at,
                "updated_at": time.time(),
                "elapsed_seconds": int(time.time() - started_at),
                "details": details or {}
            }
            if chunk_idx is not None:
                progress["current_chunk"] = chunk_idx + 1
                progress["progress"] = f"{chunk_idx + 1}/{details.get('total_chunks', '?')}"
            status_file.write_text(json.dumps(progress, indent=2))
        except Exception as e:
            print(f"Failed to write progress: {e}")

    # Parse CSV into lines
    lines = csv_text.strip().split('\n')
    if len(lines) < 2:
        return _error_response("Empty CSV")

    header = lines[0]
    rows = lines[1:]
    row_count = len(rows)

    print(f"📊 Large CSV detected: {row_count} rows. Chunking into batches of {CSV_CHUNK_SIZE}", flush=True)

    # Create chunks
    chunks = []
    for i in range(0, row_count, CSV_CHUNK_SIZE):
        chunk_rows = rows[i:i+CSV_CHUNK_SIZE]
        chunk_csv = header + '\n' + '\n'.join(chunk_rows)
        chunks.append((i, chunk_csv))

    print(f"🔀 Created {len(chunks)} chunks. Processing sequentially...", flush=True)
    write_progress("chunking_initialized", details={
        "total_rows": row_count,
        "total_chunks": len(chunks),
        "chunk_size": CSV_CHUNK_SIZE
    })

    # DIFFERENTIAL UPDATE LOGIC
    # Load existing embeddings to check if this is an update
    import hashlib
    import csv as csv_module
    from knowledge_graph.vector_store import vector_store

    old_hashes = {}
    is_differential = False

    if object_id:
        print(f"🔍 Checking for existing embeddings for object {object_id}...", flush=True)
        existing_embeddings = vector_store.get_embeddings_for_object(object_id)

        if existing_embeddings:
            print(f"📦 Found {len(existing_embeddings)} existing embeddings. Enabling differential update mode.", flush=True)
            is_differential = True

            # Build hash map from existing embeddings
            for emb in existing_embeddings:
                meta = emb["metadata"]
                idx = meta.get("embedding_index")
                row_hash = meta.get("row_hash")
                if idx is not None and row_hash:
                    old_hashes[idx] = row_hash

            print(f"📋 Loaded {len(old_hashes)} row hashes from existing embeddings", flush=True)
        else:
            print(f"🆕 No existing embeddings found. Full processing mode.", flush=True)

    # Calculate hashes for all current rows
    current_row_hashes = {}
    reader = csv_module.DictReader(lines)
    for idx, row in enumerate(reader):
        # Create stable hash from row data (sorted keys for consistency)
        row_json = json.dumps(row, sort_keys=True)
        row_hash = hashlib.md5(row_json.encode()).hexdigest()
        current_row_hashes[idx] = row_hash

    # Identify changed/new/deleted rows
    changed_row_indices = set()
    deleted_row_indices = set()
    unchanged_count = 0

    if is_differential:
        # Find changed and new rows
        for idx, new_hash in current_row_hashes.items():
            if idx not in old_hashes:
                # New row
                changed_row_indices.add(idx)
            elif old_hashes[idx] != new_hash:
                # Changed row
                changed_row_indices.add(idx)
            else:
                # Unchanged row
                unchanged_count += 1

        # Find deleted rows
        deleted_row_indices = set(old_hashes.keys()) - set(current_row_hashes.keys())

        print(f"📊 Differential analysis:", flush=True)
        print(f"   • {len(changed_row_indices)} rows changed/new", flush=True)
        print(f"   • {unchanged_count} rows unchanged", flush=True)
        print(f"   • {len(deleted_row_indices)} rows deleted", flush=True)

        write_progress("differential_analysis", details={
            "mode": "differential",
            "changed_rows": len(changed_row_indices),
            "unchanged_rows": unchanged_count,
            "deleted_rows": len(deleted_row_indices)
        })

        # Delete embeddings for removed rows
        if deleted_row_indices:
            for idx in deleted_row_indices:
                chroma_id = f"obj_{object_id}_{idx}"
                try:
                    vector_store.collection.delete(ids=[chroma_id])
                    print(f"  🗑️  Deleted embedding for removed row {idx}", flush=True)
                except Exception as e:
                    print(f"  ⚠️  Could not delete embedding {chroma_id}: {e}", flush=True)
    else:
        # Full processing mode - all rows are "changed"
        changed_row_indices = set(range(row_count))
        print(f"📊 Full processing mode: {len(changed_row_indices)} rows to process", flush=True)

    # Process chunks sequentially to avoid rate limits
    # Only process chunks that contain changed rows
    all_embeddings = []
    quality_scores = []
    chunks_processed = 0
    chunks_skipped = 0

    for chunk_idx, (start_row, chunk_csv) in enumerate(chunks):
        chunk_row_count = len(chunk_csv.split('\n')) - 1  # Exclude header
        chunk_end_row = start_row + chunk_row_count

        # Check if this chunk contains any changed rows
        chunk_has_changes = any(
            start_row <= idx < chunk_end_row
            for idx in changed_row_indices
        )

        if not chunk_has_changes and is_differential:
            print(f"  ⏭️  Skipping chunk {chunk_idx + 1}/{len(chunks)} (no changes in rows {start_row + 1}-{chunk_end_row})", flush=True)
            chunks_skipped += 1
            continue

        chunks_processed += 1
        print(f"  Processing chunk {chunk_idx + 1}/{len(chunks)} (rows {start_row + 1}-{chunk_end_row})", flush=True)

        write_progress("processing_chunk", chunk_idx, {
            "total_chunks": len(chunks),
            "chunks_processed": chunks_processed,
            "chunks_skipped": chunks_skipped,
            "rows_in_chunk": chunk_row_count,
            "embeddings_so_far": len(all_embeddings)
        })

        result = await _analyze_csv_chunk(chunk_csv, chunk_idx, context_info)

        embeddings = result.get('embeddingsList', [])

        # Add row_hash to each embedding's metadata
        for emb in embeddings:
            row_idx = emb.get("metadata", {}).get("embedding_index")
            if row_idx is not None and row_idx in current_row_hashes:
                if "metadata" not in emb:
                    emb["metadata"] = {}
                emb["metadata"]["row_hash"] = current_row_hashes[row_idx]

        all_embeddings.extend(embeddings)
        quality_scores.append(result.get('quality_score', 0))

        print(f"  ✓ Chunk {chunk_idx + 1} completed: {len(embeddings)} embeddings extracted", flush=True)

        write_progress("chunk_completed", chunk_idx, {
            "total_chunks": len(chunks),
            "chunks_processed": chunks_processed,
            "chunks_skipped": chunks_skipped,
            "embeddings_in_chunk": len(embeddings),
            "total_embeddings": len(all_embeddings)
        })

    # Calculate average quality
    avg_quality = int(sum(quality_scores) / len(quality_scores)) if quality_scores else 5

    # Prepare differential summary
    processing_mode = "differential" if is_differential else "full"
    summary_msg = f"✅ {processing_mode.title()} processing complete: {len(all_embeddings)} total embeddings from {row_count} rows"

    if is_differential:
        summary_msg += f" ({chunks_skipped} chunks skipped)"

    print(summary_msg, flush=True)
    write_progress("merging_results", details={
        "total_embeddings": len(all_embeddings),
        "total_rows": row_count,
        "avg_quality": avg_quality,
        "processing_mode": processing_mode,
        "chunks_processed": chunks_processed,
        "chunks_skipped": chunks_skipped
    })

    # Build differential metadata
    differential_info = {}
    if is_differential:
        differential_info = {
            "mode": "differential",
            "changed_rows": len(changed_row_indices),
            "unchanged_rows": unchanged_count,
            "deleted_rows": len(deleted_row_indices),
            "chunks_processed": chunks_processed,
            "chunks_skipped": chunks_skipped
        }
    else:
        differential_info = {
            "mode": "full",
            "total_rows_processed": row_count
        }

    return {
        "category": "document",
        "danger_potential": 1,
        "safety_info": {
            "isSafe": True,
            "confidence": 1.0,
            "reasoning": f"CSV data processed in {processing_mode} mode",
            "flags": []
        },
        "ai_title": "CSV Data Import",
        "ai_subtitle": f"📊 Structured data processed ({processing_mode} mode)",
        "ai_tags": ["csv", "data", "import"],
        "ai_collections": ["Data"],
        "extracted_tags": {"keywords": ["csv", "structured_data"]},
        "embedding_info": {
            "embeddingText": f"CSV file with {row_count} rows processed in {processing_mode} mode",
            "searchableFields": ["data"],
            "metadata": {
                "total_rows": row_count,
                "chunks_total": len(chunks),
                "chunks_processed": chunks_processed,
                "chunks_skipped": chunks_skipped if is_differential else 0,
                "chunk_size": CSV_CHUNK_SIZE,
                "differential_update": differential_info
            },
            "embeddingQuality": {
                "quality_score": avg_quality,
                "needs_review": False,
                "issues": [],
                "recommendation": "auto_embed"
            },
            "embeddingsList": all_embeddings
        },
        "mode": "chunked",
        "prompt": f"Chunked CSV processing: {len(chunks)} chunks",
        "ai_response": f"Processed {len(all_embeddings)} embeddings from {row_count} rows"
    }


async def _analyze_vision_comprehensive(
    data: bytes,
    context: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    VISION MODE: Comprehensive image analysis with product detection and visual intelligence.

    This function performs deep visual analysis including:
    - Product detection and property estimation
    - Color analysis and visual harmony
    - Composition and layout intelligence
    - Semantic properties for catalog organization

    Args:
        data: Image content as bytes
        context: Optional contextual metadata

    Returns:
        Dict with comprehensive vision analysis and embedding data
    """
    context_info = build_context_info(context)
    prompt_text = VISION_ANALYSIS_PROMPT.format(context_info=context_info)

    # Encode image
    encoded_image = base64.b64encode(data).decode("utf-8")
    images_list = [encoded_image]

    # Call Gemini Vision API
    payload = {"prompt": {"text": prompt_text, "images": images_list}}
    headers = {"X-API-KEY": INTERNAL_API_KEY, "Content-Type": "application/json"}

    async with httpx.AsyncClient() as client:
        try:
            print("🎨 Starting comprehensive vision analysis...", flush=True)
            response = await client.post(
                f"{API_BASE_URL}/ai/gemini",
                json=payload,
                headers=headers,
                timeout=60.0  # Vision analysis should be faster than CSV
            )
            response.raise_for_status()

            ai_response_str = response.json().get("message", "{}")
            ai_response_str = _clean_json_response(ai_response_str)
            analysis_result = json.loads(ai_response_str)

            # Extract all components
            safety_check = analysis_result.get("safetyCheck", {})
            classification = analysis_result.get("classification", {})
            product_analysis = analysis_result.get("productAnalysis", {})
            visual_analysis = analysis_result.get("visualAnalysis", {})
            layout_intelligence = analysis_result.get("layoutIntelligence", {})
            semantic_props = analysis_result.get("semanticProperties", {})
            tech_metadata = analysis_result.get("technicalMetadata", {})
            media_analysis = analysis_result.get("mediaAnalysis", {})
            embedding_info = analysis_result.get("embeddingInfo", {})

            # Merge all visual intelligence into embedding metadata
            if "metadata" not in embedding_info:
                embedding_info["metadata"] = {}

            embedding_info["metadata"].update({
                "product_analysis": product_analysis,
                "visual_analysis": visual_analysis,
                "layout_intelligence": layout_intelligence,
                "semantic_properties": semantic_props,
                "technical_metadata": tech_metadata,
                "is_product": classification.get("isProduct", False)
            })

            # Build comprehensive extracted_tags
            extracted_tags = {
                "keywords": semantic_props.get("keywords", []),
                "colors": product_analysis.get("colors", []),
                "materials": product_analysis.get("materials", []),
                "visual_harmony_tags": layout_intelligence.get("visualHarmonyTags", [])
            }

            print(f"✅ Vision analysis complete. Product: {classification.get('isProduct', False)}", flush=True)

            return {
                "category": classification.get("category", "unknown"),
                "danger_potential": classification.get("dangerPotential", 1),
                "safety_info": {
                    "isSafe": safety_check.get("isSafe", True),
                    "confidence": safety_check.get("confidence", 1.0),
                    "reasoning": safety_check.get("reasoning", ""),
                    "flags": safety_check.get("flags", [])
                },
                "ai_title": media_analysis.get("suggestedTitle"),
                "ai_subtitle": media_analysis.get("suggestedSubtitle"),
                "ai_tags": media_analysis.get("tags", []),
                "ai_collections": media_analysis.get("collectionSuggestions", []),
                "extracted_tags": extracted_tags,
                "embedding_info": embedding_info,
                "mode": "vision_comprehensive",
                "prompt": prompt_text,
                "ai_response": ai_response_str
            }

        except Exception as e:
            print(f"❌ Vision analysis failed: {e}", flush=True)
            import traceback
            traceback.print_exc()
            return _error_response(f"Vision analysis failed: {e}")


async def analyze_content(
    data: bytes,
    mime_type: str,
    context: Optional[Dict[str, Any]] = None,
    object_id: Optional[int] = None
) -> Dict[str, Any]:
    """
    Analyze content using AI.

    Supports multiple modes:
    - VISION (for images): Comprehensive product/visual analysis
    - UNIFIED (default): Single AI request for all analysis
    - SPLIT: Separate requests for safety and embedding

    Args:
        data: File content as bytes
        mime_type: MIME type of the file
        context: Optional contextual metadata
        object_id: Optional object ID for progress tracking

    Returns:
        Dict with safety_info, category, tags, collections, extracted_tags, embedding_info
    """
    # Use comprehensive vision analysis for images
    if mime_type.startswith("image/"):
        print(f"🎨 Detected image upload - using comprehensive vision analysis", flush=True)
        return await _analyze_vision_comprehensive(data, context)

    # For non-images, use existing modes
    if ai_config.is_split():
        return await _analyze_split_mode(data, mime_type, context, object_id)
    else:
        return await _analyze_unified_mode(data, mime_type, context, object_id)


async def _analyze_unified_mode(
    data: bytes,
    mime_type: str,
    context: Optional[Dict[str, Any]] = None,
    object_id: Optional[int] = None
) -> Dict[str, Any]:
    """UNIFIED MODE: Single AI request for comprehensive analysis"""

    context_info = build_context_info(context)
    prompt_text = UNIFIED_PROMPT.format(context_info=context_info)

    # Prepare content
    images_list = []
    if mime_type.startswith("image/"):
        encoded_image = base64.b64encode(data).decode("utf-8")
        images_list.append(encoded_image)
    elif mime_type.startswith("text/") or mime_type == "application/csv":
        try:
            content_as_text = data.decode("utf-8")

            # Check if this is a CSV that should be chunked
            if mime_type == "application/csv":
                lines = content_as_text.strip().split('\n')
                row_count = len(lines) - 1  # Exclude header

                if row_count > CSV_MAX_ROWS_BEFORE_CHUNKING:
                    print(f"🔀 CSV has {row_count} rows, using chunked processing", flush=True)
                    return await _analyze_csv_chunked(content_as_text, context, object_id)
                else:
                    print(f"📄 CSV has {row_count} rows, using normal processing", flush=True)

            # Normal text/small CSV processing
            max_chars = 1_000_000
            if len(content_as_text) > max_chars:
                print(f"⚠️ Content truncated from {len(content_as_text)} to {max_chars} chars")
                prompt_text += f"\n\n--- CONTENT (truncated to {max_chars} chars) ---\n{content_as_text[:max_chars]}"
            else:
                prompt_text += f"\n\n--- CONTENT ---\n{content_as_text}"
        except UnicodeDecodeError:
            return _error_response("Binary file")
    elif mime_type in [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
        "application/vnd.ms-excel"  # .xls
    ]:
        try:
            # Extract Excel content as CSV-like text
            excel_text = extract_excel_as_text(data, mime_type)
            # Send full content to Gemini Pro (2M token context window)
            # Limit to 1M chars (~250K tokens) for safety
            max_chars = 1_000_000
            if len(excel_text) > max_chars:
                print(f"⚠️ Excel content truncated from {len(excel_text)} to {max_chars} chars")
                prompt_text += f"\n\n--- EXCEL CONTENT (truncated to {max_chars} chars) ---\n{excel_text[:max_chars]}"
            else:
                prompt_text += f"\n\n--- EXCEL CONTENT ---\n{excel_text}"
            print(f"📊 Extracted {len(excel_text)} chars from Excel file")
        except Exception as e:
            print(f"⚠️ Excel parsing failed: {e}")
            return _error_response(f"Excel parsing failed: {e}")
    else:
        return _error_response("Unsupported type")

    # Call Gemini API
    payload = {"prompt": {"text": prompt_text, "images": images_list}}
    headers = {"X-API-KEY": INTERNAL_API_KEY, "Content-Type": "application/json"}

    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{API_BASE_URL}/ai/gemini",
                json=payload,
                headers=headers,
                timeout=1200.0  # 20 minutes for large CSV files with Gemini Pro
            )
            response.raise_for_status()

            ai_response_str = response.json().get("message", "{}")
            ai_response_str = _clean_json_response(ai_response_str)
            analysis_result = json.loads(ai_response_str)

            # Extract components
            safety_check = analysis_result.get("safetyCheck", {})
            classification = analysis_result.get("classification", {})
            media_analysis = analysis_result.get("mediaAnalysis", {})
            extracted_tags = analysis_result.get("extractedTags", {})
            embedding_info = analysis_result.get("embeddingInfo", {})

            return {
                "category": classification.get("category", "unknown"),
                "danger_potential": classification.get("dangerPotential", 1),
                "safety_info": {
                    "isSafe": safety_check.get("isSafe", True),
                    "confidence": safety_check.get("confidence", 1.0),
                    "reasoning": safety_check.get("reasoning", ""),
                    "flags": safety_check.get("flags", [])
                },
                "ai_title": media_analysis.get("suggestedTitle"),
                "ai_subtitle": media_analysis.get("suggestedSubtitle"),
                "ai_tags": media_analysis.get("tags", []),
                "ai_collections": media_analysis.get("collectionSuggestions", []),
                "extracted_tags": extracted_tags,
                "embedding_info": embedding_info,
                "mode": str(ai_config.mode),
                "prompt": prompt_text,
                "ai_response": ai_response_str
            }

        except Exception as e:
            print(f"AI analysis failed: {e}")
            import traceback
            traceback.print_exc()
            return _error_response(f"Analysis failed: {e}")


async def _analyze_split_mode(
    data: bytes,
    mime_type: str,
    context: Optional[Dict[str, Any]] = None,
    object_id: Optional[int] = None
) -> Dict[str, Any]:
    """SPLIT MODE: Separate AI requests for safety and embedding"""

    print(f"🔄 SPLIT MODE: Safety={ai_config.safety_model}, Embedding={ai_config.embedding_model}")

    # Prepare content
    images_list = []
    text_content = None

    if mime_type.startswith("image/"):
        encoded_image = base64.b64encode(data).decode("utf-8")
        images_list.append(encoded_image)
    elif mime_type.startswith("text/") or mime_type == "application/csv":
        try:
            full_text = data.decode("utf-8")

            # Check if this is a large CSV that should be chunked
            if mime_type == "application/csv":
                lines = full_text.strip().split('\n')
                row_count = len(lines) - 1  # Exclude header

                if row_count > CSV_MAX_ROWS_BEFORE_CHUNKING:
                    print(f"🔀 CSV has {row_count} rows, using chunked processing (SPLIT mode)")
                    return await _analyze_csv_chunked(full_text, context, object_id)
                else:
                    print(f"📄 CSV has {row_count} rows, using normal processing (SPLIT mode)")

            # Normal text/small CSV processing
            max_chars = 1_000_000
            if len(full_text) > max_chars:
                print(f"⚠️ Text content truncated from {len(full_text)} to {max_chars} chars")
                text_content = full_text[:max_chars]
            else:
                text_content = full_text
        except UnicodeDecodeError:
            return _error_response("Binary file")
    elif mime_type in [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
        "application/vnd.ms-excel"  # .xls
    ]:
        try:
            # Extract Excel content as CSV-like text
            excel_text = extract_excel_as_text(data, mime_type)
            # Send full content to Gemini Pro (2M token context window)
            # Limit to 1M chars (~250K tokens) for safety
            max_chars = 1_000_000
            if len(excel_text) > max_chars:
                print(f"⚠️ Excel content truncated from {len(excel_text)} to {max_chars} chars")
                text_content = excel_text[:max_chars]
            else:
                text_content = excel_text
            print(f"📊 Extracted {len(excel_text)} chars from Excel file")
        except Exception as e:
            print(f"⚠️ Excel parsing failed: {e}")
            return _error_response(f"Excel parsing failed: {e}")
    else:
        return _error_response("Unsupported type")

    context_info = build_context_info(context)

    # STEP 1: Safety Check
    safety_result = await _run_safety_check(images_list, text_content, context_info)

    # STEP 2: Embedding Generation
    embedding_result = await _run_embedding_generation(images_list, text_content, context_info)

    # Merge results with combined prompts and responses
    combined_prompt = f"""=== SPLIT MODE ===
Safety Model: {ai_config.safety_model}
Embedding Model: {ai_config.embedding_model}

--- SAFETY CHECK PROMPT ---
{safety_result.get('prompt', 'N/A')}

--- EMBEDDING GENERATION PROMPT ---
{embedding_result.get('prompt', 'N/A')}"""

    combined_response = f"""=== SAFETY CHECK RESPONSE ===
{safety_result.get('ai_response', 'N/A')}

=== EMBEDDING GENERATION RESPONSE ===
{embedding_result.get('ai_response', 'N/A')}"""

    return {
        "category": safety_result.get("category", "unknown"),
        "danger_potential": safety_result.get("danger_potential", 1),
        "safety_info": safety_result.get("safety_info", {}),
        "ai_title": embedding_result.get("ai_title"),
        "ai_subtitle": embedding_result.get("ai_subtitle"),
        "ai_tags": embedding_result.get("ai_tags", []),
        "ai_collections": embedding_result.get("ai_collections", []),
        "extracted_tags": embedding_result.get("extracted_tags", {}),
        "embedding_info": embedding_result.get("embedding_info", {}),
        "mode": str(ai_config.mode),
        "prompt": combined_prompt,
        "ai_response": combined_response
    }


async def _run_safety_check(
    images_list: list,
    text_content: Optional[str],
    context_info: str
) -> Dict[str, Any]:
    """Run safety check only"""

    prompt = SAFETY_PROMPT.format(context_info=context_info)
    if text_content:
        prompt += f"\n\n--- CONTENT ---\n{text_content}"

    payload = {"prompt": {"text": prompt, "images": images_list}}
    headers = {"X-API-KEY": INTERNAL_API_KEY, "Content-Type": "application/json"}

    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{API_BASE_URL}/ai/gemini",
                json=payload,
                headers=headers,
                timeout=30.0
            )
            response.raise_for_status()

            ai_response_str = response.json().get("message", "{}")
            ai_response_str = _clean_json_response(ai_response_str)
            result = json.loads(ai_response_str)

            safety_check = result.get("safetyCheck", {})
            classification = result.get("classification", {})

            return {
                "category": classification.get("category", "unknown"),
                "danger_potential": classification.get("dangerPotential", 1),
                "safety_info": {
                    "isSafe": safety_check.get("isSafe", True),
                    "confidence": safety_check.get("confidence", 1.0),
                    "reasoning": safety_check.get("reasoning", ""),
                    "flags": safety_check.get("flags", [])
                },
                "prompt": prompt,
                "ai_response": ai_response_str
            }
        except Exception as e:
            print(f"Safety check failed: {e}")
            return _error_response(f"Safety check failed: {e}")


async def _run_embedding_generation(
    images_list: list,
    text_content: Optional[str],
    context_info: str
) -> Dict[str, Any]:
    """Run embedding generation"""

    prompt = EMBEDDING_PROMPT.format(context_info=context_info)
    if text_content:
        prompt += f"\n\n--- CONTENT ---\n{text_content}"

    payload = {"prompt": {"text": prompt, "images": images_list}}
    headers = {"X-API-KEY": INTERNAL_API_KEY, "Content-Type": "application/json"}

    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{API_BASE_URL}/ai/gemini",
                json=payload,
                headers=headers,
                timeout=1200.0  # 20 minutes for large CSV files with Gemini Pro
            )
            response.raise_for_status()

            ai_response_str = response.json().get("message", "{}")
            ai_response_str = _clean_json_response(ai_response_str)
            result = json.loads(ai_response_str)

            media_analysis = result.get("mediaAnalysis", {})
            extracted_tags = result.get("extractedTags", {})
            embedding_info = result.get("embeddingInfo", {})

            return {
                "ai_title": media_analysis.get("suggestedTitle"),
                "ai_subtitle": media_analysis.get("suggestedSubtitle"),
                "ai_tags": media_analysis.get("tags", []),
                "ai_collections": media_analysis.get("collectionSuggestions", []),
                "extracted_tags": extracted_tags,
                "embedding_info": embedding_info,
                "prompt": prompt,
                "ai_response": ai_response_str
            }
        except Exception as e:
            print(f"Embedding generation failed: {e}")
            return {
                "ai_title": None,
                "ai_subtitle": None,
                "ai_tags": [],
                "ai_collections": [],
                "extracted_tags": {},
                "embedding_info": {}
            }


def _clean_json_response(response: str) -> str:
    """Clean JSON markers from AI response"""
    response = response.strip()
    if response.startswith("```json"):
        response = response[7:].strip()
    if response.startswith("```"):
        response = response[3:].strip()
    if response.endswith("```"):
        response = response[:-3].strip()
    return response


def _error_response(reason: str) -> Dict[str, Any]:
    """Return error response"""
    return {
        "category": "error",
        "danger_potential": 1,
        "safety_info": {"isSafe": True, "confidence": 0.5, "reasoning": reason, "flags": []}
    }
